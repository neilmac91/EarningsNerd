"""The branded Excel workbook export (excel_export_service) + its Pro-gated route.

``build_analysis_workbook`` is pure over the dataset dict, so every visual/typing contract is
verified by READING THE PRODUCED BYTES BACK with openpyxl (sheets, freeze panes, header styling,
number formats, derived-Q4 comments) — not by inspecting internals. What openpyxl cannot read
back (embedded images, native charts), the zip container proves directly: an .xlsx is a zip, so
``xl/media/*.png`` / ``xl/charts/chart*.xml`` entries pin the brand mark and the panel charts.

Route tests stub auth (the ``_resolve_current_user`` override pattern) and the company/dataset
lookups — the workbook builder itself is already covered purely above.
"""
import io
import zipfile
from datetime import datetime, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.services.excel_export_service import build_analysis_workbook

EXPORTED_AT = datetime(2026, 7, 6, tzinfo=timezone.utc)


def _usd_series(concept, label, values):
    return {
        "concept": concept,
        "label": label,
        "unit": "USD",
        "percent": False,
        "points": [{"period": key, "value": value} for key, value in values],
    }


def _dataset():
    """Annual 3-period dataset exercising every format branch and all four panels."""
    return {
        "ticker": "TST",
        "company_name": "Test & Co",
        "mode": "annual",
        "period_key": "FY2022..FY2024",
        "periods": [
            {"key": "FY2022", "fiscal_year": 2022, "fiscal_period": "FY", "period_end": "2022-12-31"},
            {"key": "FY2023", "fiscal_year": 2023, "fiscal_period": "FY", "period_end": "2023-12-31"},
            {"key": "FY2024", "fiscal_year": 2024, "fiscal_period": "FY", "period_end": "2024-12-31"},
        ],
        "series": [
            {
                "concept": "revenue", "label": "Revenue", "unit": "USD", "percent": False,
                "cagr": 0.2247, "cagr_window": "FY2022..FY2024",
                "points": [
                    {"period": "FY2022", "value": 1000.0},
                    {"period": "FY2023", "value": 1200.0, "yoy": 0.2},
                    # yoy "nm" = the sign-flip sentinel; derived = a computed-Q4 estimate.
                    {"period": "FY2024", "value": 1500.0, "yoy": "nm", "derived": True},
                ],
            },
            {
                "concept": "net_margin", "label": "Net margin", "unit": "pure", "percent": True,
                "window_pp": 4.3, "window_pp_range": "FY2022..FY2024",
                "points": [
                    {"period": "FY2022", "value": 20.0},
                    {"period": "FY2023", "value": None},
                    {"period": "FY2024", "value": 24.3},
                ],
            },
            {
                "concept": "eps_diluted", "label": "EPS (diluted)", "unit": "USD/shares",
                "percent": False,
                "points": [{"period": "FY2022", "value": 1.11}, {"period": "FY2024", "value": 1.5}],
            },
            {
                "concept": "current_ratio", "label": "Current ratio", "unit": "pure",
                "percent": False,
                "points": [{"period": "FY2022", "value": 1.5}, {"period": "FY2024", "value": 1.7}],
            },
            _usd_series("operating_cash_flow", "Operating cash flow",
                        [("FY2022", 300.0), ("FY2023", 340.0), ("FY2024", 420.0)]),
            _usd_series("free_cash_flow", "Free cash flow",
                        [("FY2022", 200.0), ("FY2023", 250.0), ("FY2024", 310.0)]),
            _usd_series("net_income", "Net income",
                        [("FY2022", 180.0), ("FY2023", 220.0), ("FY2024", 364.0)]),
            _usd_series("long_term_debt", "Long-term debt",
                        [("FY2022", 500.0), ("FY2023", 480.0), ("FY2024", 450.0)]),
            _usd_series("shareholders_equity", "Shareholders' equity",
                        [("FY2022", 900.0), ("FY2023", 1100.0), ("FY2024", 1400.0)]),
            _usd_series("cash_and_equivalents", "Cash & equivalents",
                        [("FY2022", 250.0), ("FY2023", 260.0), ("FY2024", 400.0)]),
        ],
    }


@pytest.fixture(scope="module")
def workbook_bytes():
    return build_analysis_workbook(_dataset(), exported_at=EXPORTED_AT)


@pytest.fixture(scope="module")
def workbook(workbook_bytes):
    return load_workbook(io.BytesIO(workbook_bytes))


class TestWorkbookStructure:
    def test_bytes_are_a_zip_workbook(self, workbook_bytes):
        assert workbook_bytes[:2] == b"PK"

    def test_sheet_order_overview_metrics_then_panels(self, workbook):
        assert workbook.sheetnames == [
            "Overview",
            "Metrics",
            "Revenue & growth",
            "Margins",
            "Cash generation",
            "Balance sheet",
        ]

    def test_every_sheet_carries_the_sage_tab(self, workbook):
        for name in workbook.sheetnames:
            tab = workbook[name].sheet_properties.tabColor
            assert tab is not None and str(tab.rgb).endswith("4F7A63"), name

    def test_absent_concept_panels_are_skipped_and_nii_names_the_bar_panel(self):
        nii_only = {
            "ticker": "BNK", "company_name": "Bank Co", "mode": "annual",
            "period_key": "FY2023..FY2024",
            "periods": [
                {"key": "FY2023", "fiscal_year": 2023, "fiscal_period": "FY", "period_end": "2023-12-31"},
                {"key": "FY2024", "fiscal_year": 2024, "fiscal_period": "FY", "period_end": "2024-12-31"},
            ],
            "series": [
                {
                    "concept": "net_interest_income", "label": "Net interest income",
                    "unit": "USD", "percent": False,
                    "points": [{"period": "FY2023", "value": 90.0}, {"period": "FY2024", "value": 110.0}],
                },
                _usd_series("net_income", "Net income", [("FY2023", 40.0), ("FY2024", 55.0)]),
            ],
        }
        wb = load_workbook(io.BytesIO(build_analysis_workbook(nii_only, exported_at=EXPORTED_AT)))
        # No margins / balance-sheet concepts → those panels collapse, like the frontend's.
        assert wb.sheetnames == [
            "Overview", "Metrics", "Net interest income & growth", "Cash generation",
        ]


class TestOverviewSheet:
    def _texts(self, workbook):
        ws = workbook["Overview"]
        return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]

    def test_title_provenance_and_terms(self, workbook):
        texts = self._texts(workbook)
        assert any("Test & Co — Multi-Period Analysis" in t for t in texts)
        assert "TST" in texts
        assert any("FY2022..FY2024 (3 periods)" in t for t in texts)
        assert any("July 06, 2026" in t for t in texts)
        assert any("earningsnerd.io/terms" in t for t in texts)
        assert any("not" in t and "investment" in t for t in texts)  # the disclaimer paragraph

    def test_conventions_are_stated(self, workbook):
        texts = self._texts(workbook)
        assert any("true Excel percentages" in t for t in texts)
        assert any("Computed Q4" in t or "computed Q4" in t for t in texts)  # † legend (derived present)

    def test_brand_mark_png_is_embedded(self, workbook_bytes):
        # openpyxl does not read images back — the zip container is the proof.
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            assert any(
                name.startswith("xl/media/") and name.endswith(".png")
                for name in archive.namelist()
            )


class TestMetricsSheet:
    def test_freeze_panes_and_header_styling(self, workbook):
        ws = workbook["Metrics"]
        assert ws.freeze_panes == "B2"
        header = ws["A1"]
        assert header.value == "Metric"
        assert str(header.fill.start_color.rgb).endswith("ECF2EE")
        assert header.font.bold and str(header.font.color.rgb).endswith("3C6650")
        assert ws["D1"].value == "FY2022"

    def test_number_formats_per_unit(self, workbook):
        ws = workbook["Metrics"]
        # Row order == dataset series order: 2=revenue, 3=net_margin, 4=eps, 5=current_ratio.
        assert ws["D2"].value == 1000.0 and ws["D2"].number_format == "#,##0"
        # Percent series are TRUE Excel percents: stored ÷100, formatted 0.0%.
        assert ws["D3"].value == pytest.approx(0.20) and ws["D3"].number_format == "0.0%"
        assert ws["E3"].value is None  # missing point stays empty
        assert ws["D4"].value == 1.11 and ws["D4"].number_format == "#,##0.00"
        assert ws["D5"].value == 1.5 and ws["D5"].number_format == '0.00"x"'

    def test_derived_q4_gets_a_comment_never_a_fill(self, workbook):
        ws = workbook["Metrics"]
        derived = ws["F2"]  # revenue FY2024
        assert derived.comment is not None and "Computed Q4" in derived.comment.text
        assert derived.fill.patternType is None  # no fill — comments only
        assert ws["D2"].comment is None  # non-derived neighbour carries none

    def test_annual_window_figures(self, workbook):
        ws = workbook["Metrics"]
        assert ws["G1"].value == "Window growth" and ws["H1"].value == "Window"
        assert ws["G2"].value == pytest.approx(0.2247) and ws["G2"].number_format == "0.0%"
        assert ws["H2"].value == "FY2022..FY2024"
        # Percent series get the pp window figure instead of CAGR.
        assert ws["G3"].value == pytest.approx(4.3) and ws["G3"].number_format == '0.0" pp"'

    def test_quarterly_dataset_omits_window_columns(self):
        quarterly = _dataset()
        quarterly["mode"] = "quarterly"
        for series in quarterly["series"]:
            series.pop("cagr", None)
            series.pop("cagr_window", None)
            series.pop("window_pp", None)
            series.pop("window_pp_range", None)
        wb = load_workbook(io.BytesIO(build_analysis_workbook(quarterly, exported_at=EXPORTED_AT)))
        ws = wb["Metrics"]
        headers = [c.value for c in ws[1]]
        assert "Window growth" not in headers and "Window" not in headers


class TestPanelSheets:
    def test_top_line_sheet_ships_growth_as_a_column(self, workbook):
        ws = workbook["Revenue & growth"]
        assert [c.value for c in ws[1]] == ["Period", "Revenue", "YoY growth"]
        assert ws["B2"].value == 1000.0 and ws["B2"].number_format == "#,##0"
        assert ws["C2"].value is None  # first period has no YoY
        assert ws["C3"].value == pytest.approx(0.2) and ws["C3"].number_format == "0.0%"
        assert ws["C4"].value == "n/m"  # the sign-flip sentinel is text, not a fake percentage

    def test_margins_sheet_writes_true_percents(self, workbook):
        ws = workbook["Margins"]
        # Only net_margin exists in the fixture — absent concepts drop their columns.
        assert [c.value for c in ws[1]] == ["Period", "Net"]
        assert ws["B2"].value == pytest.approx(0.20) and ws["B2"].number_format == "0.0%"

    def test_formula_injection_is_neutralized(self):
        """SEC-sourced strings (company name, ticker, XBRL units) must never become live Excel
        formulas — openpyxl persists any string cell starting with "=" as data_type 'f'. The
        guard forces the xlsx type attribute back to string, so the value displays VERBATIM
        (no apostrophe mutation) yet Excel never evaluates it."""
        hostile = _dataset()
        hostile["company_name"] = '=HYPERLINK("http://evil.example","x")'
        hostile["ticker"] = "=T"
        hostile["series"][2]["unit"] = "=cmd/shares"  # eps_diluted row
        data = build_analysis_workbook(hostile, exported_at=EXPORTED_AT)
        wb = load_workbook(io.BytesIO(data))

        overview = wb["Overview"]
        title = overview["A7"]
        assert title.data_type == "s" and title.value.startswith("=HYPERLINK")
        ticker_cell = overview["B10"]
        assert ticker_cell.data_type == "s" and ticker_cell.value == "=T"
        metrics = wb["Metrics"]
        unit_cell = metrics["C4"]  # eps_diluted unit
        assert unit_cell.data_type == "s" and unit_cell.value == "=cmd/shares"

        # Structural proof: the workbook writes NO formulas anywhere — any <f> element in a
        # worksheet part would mean an injected value slipped through as a live formula.
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            for name in archive.namelist():
                if name.startswith("xl/worksheets/"):
                    assert b"<f>" not in archive.read(name), name

    def test_native_charts_use_the_design_system_series_colors(self, workbook_bytes):
        # openpyxl does not read charts back — assert on the chart parts in the zip container.
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            chart_parts = [n for n in archive.namelist() if n.startswith("xl/charts/chart")]
            assert len(chart_parts) == 4  # one per panel
            chart_xml = "".join(archive.read(n).decode("utf-8") for n in chart_parts)
        assert "3E8E84" in chart_xml  # series 1 (teal) — bar + first lines
        assert "B8812F" in chart_xml  # series 2 (honey)
        assert "5B7CC0" in chart_xml  # series 3 (cornflower)


@pytest.mark.requires_db
class TestXlsxRoute:
    """Route contract: Pro `can_export` gate, headers, and the xlsx payload."""

    @pytest.fixture()
    def client(self):
        from fastapi.testclient import TestClient

        from main import app

        with TestClient(app) as test_client:
            yield test_client
            app.dependency_overrides.clear()

    def _as_user(self, is_pro):
        from app.dependencies import _resolve_current_user
        from main import app

        user = SimpleNamespace(id=987_654, is_pro=is_pro, subscription=None)
        app.dependency_overrides[_resolve_current_user] = lambda: user
        return user

    def test_unauthenticated_is_401(self, client):
        response = client.post(
            "/api/analysis/TST/export/xlsx",
            json={"mode": "annual", "start_period": "FY2022", "end_period": "FY2024"},
        )
        assert response.status_code == 401

    def test_free_user_is_403_with_upsell(self, client):
        self._as_user(is_pro=False)
        response = client.post(
            "/api/analysis/TST/export/xlsx",
            json={"mode": "annual", "start_period": "FY2022", "end_period": "FY2024"},
        )
        assert response.status_code == 403
        assert "Excel export" in response.json()["detail"]
        assert "Upgrade to Pro" in response.json()["detail"]

    def test_pro_user_downloads_the_workbook(self, client, monkeypatch):
        import app.routers.analysis as analysis_router

        self._as_user(is_pro=True)
        monkeypatch.setattr(
            analysis_router, "_get_company", lambda db, ticker: SimpleNamespace(ticker="TST")
        )
        monkeypatch.setattr(
            analysis_router.trend_analysis_service,
            "build_dataset",
            lambda db, company, mode, start, end: _dataset(),
        )
        response = client.post(
            "/api/analysis/TST/export/xlsx",
            json={"mode": "annual", "start_period": "FY2022", "end_period": "FY2024"},
        )
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert (
            response.headers["content-disposition"]
            == 'attachment; filename="TST_FY2022-FY2024_annual-metrics.xlsx"'
        )
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.sheetnames[0] == "Overview"

    def test_bad_range_maps_to_400(self, client, monkeypatch):
        import app.routers.analysis as analysis_router

        self._as_user(is_pro=True)
        monkeypatch.setattr(
            analysis_router, "_get_company", lambda db, ticker: SimpleNamespace(ticker="TST")
        )

        def _raises(db, company, mode, start, end):
            raise ValueError("No data available for the selected period range.")

        monkeypatch.setattr(analysis_router.trend_analysis_service, "build_dataset", _raises)
        response = client.post(
            "/api/analysis/TST/export/xlsx",
            json={"mode": "annual", "start_period": "FY1900", "end_period": "FY1901"},
        )
        assert response.status_code == 400
