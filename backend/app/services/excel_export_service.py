"""Branded Excel workbook export for Multi-Period Analysis (owner decision D1).

One workbook replaces the old per-dataset CSV download: an **Overview** sheet (brand mark +
provenance + conventions + disclaimer), the full **Metrics** grid (typed cells, real Excel number
formats, derived-Q4 cell comments), and **one data sheet per chart panel** with a native chart —
a single download carrying everything the four on-page trend panels show.

``build_analysis_workbook`` is a PURE function over the deterministic dataset dict
(``trend_analysis_service.build_dataset``) — no DB, no I/O beyond the committed brand PNG — so
tests read the produced bytes straight back with openpyxl.

Conventions (also stated on the Overview sheet, where users see them):

- Monetary values are RAW dollars at full precision, formatted ``#,##0`` — the compact "$24.9B"
  rendering belongs to the PDF and on-page table; the workbook is the machine-usable surface.
- Percent series (margins, YoY growth) are written as TRUE Excel percentages: the cell stores
  the fraction (0.558) and formats as ``0.0%``. Dataset margin values arrive ×100 and are divided
  back down on write; growth ratios arrive as fractions and are written as-is.
- Derived Q4 estimates carry a cell Comment, never a fill (a fill reads as data emphasis).
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.pdf_branding import PALETTE
from app.utils.datetimes import utcnow

# 256px-wide sage mark on transparent, rasterized by frontend/scripts/generate-brand-assets.mjs
# (openpyxl can't rasterize SVG; the PNG is committed so the backend needs no SVG toolchain).
_MARK_PNG = Path(__file__).resolve().parent.parent / "assets" / "brand" / "earningsnerd-mark.png"

# frontend/components/ui/Chart.tsx CHART_SERIES — positional 1→N, never re-sorted, so workbook
# charts pair series→color exactly like the on-page panels (openpyxl hex carries no '#').
CHART_SERIES: tuple[str, ...] = ("3E8E84", "B8812F", "5B7CC0", "CF7159", "6E7E9C", "8B7BC0")

_TAB_COLOR = PALETTE["brand"].lstrip("#")
_HEADER_FILL = PatternFill("solid", start_color=PALETTE["brand_weak"].lstrip("#"))
_HEADER_FONT = Font(bold=True, color=PALETTE["brand_strong"].lstrip("#"), size=10)
_INK_FONT = Font(color=PALETTE["ink"].lstrip("#"))
_LABEL_FONT = Font(bold=True, color=PALETTE["brand_strong"].lstrip("#"))
_FOOTNOTE_FONT = Font(color=PALETTE["ink_tertiary"].lstrip("#"), size=9)
_HAIRLINE = Side(style="thin", color=PALETTE["border"].lstrip("#"))
_CELL_BORDER = Border(left=_HAIRLINE, right=_HAIRLINE, top=_HAIRLINE, bottom=_HAIRLINE)

_DERIVED_COMMENT = (
    "Computed Q4 — derived from the annual report: full year minus the reported year-to-date "
    "quarters (EPS: Q4 net income ÷ Q4 weighted shares). This estimate does not appear in the "
    "filings themselves."
)

# The three line panels from TrendCharts.tsx (the bar panel is built separately — its top-line
# concept is dataset-dependent). Panels whose concepts are all absent are skipped, mirroring the
# frontend's collapsing panels (a bank has no gross margin).
_LINE_PANELS: tuple[tuple[str, tuple[tuple[str, str], ...]], ...] = (
    (
        "Margins",
        (("gross_margin", "Gross"), ("operating_margin", "Operating"), ("net_margin", "Net")),
    ),
    (
        "Cash generation",
        (
            ("operating_cash_flow", "Operating CF"),
            ("free_cash_flow", "Free cash flow"),
            ("net_income", "Net income"),
        ),
    ),
    (
        "Balance sheet",
        (
            ("long_term_debt", "Long-term debt"),
            ("shareholders_equity", "Equity"),
            ("cash_and_equivalents", "Cash"),
        ),
    ),
)

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(title: str) -> str:
    """Excel sheet names: 31 chars max, no []:*?/\\ characters."""
    return _INVALID_SHEET_CHARS.sub("-", title)[:31]


def _number_format(unit: Any, percent: bool) -> str:
    if percent:
        return "0.0%"
    if unit == "pure":
        return '0.00"x"'
    if isinstance(unit, str) and unit.endswith("/shares"):
        return "#,##0.00"
    return "#,##0"


def _write_point(
    ws: Worksheet, row: int, col: int, point: dict[str, Any], unit: Any, percent: bool
) -> None:
    """One dataset point → one typed, formatted, bordered cell (comment for derived Q4)."""
    cell = ws.cell(row=row, column=col)
    cell.border = _CELL_BORDER
    value = point.get("value")
    if value is None:
        return
    cell.value = value / 100 if percent else value
    cell.number_format = _number_format(unit, percent)
    if point.get("derived"):
        cell.comment = Comment(_DERIVED_COMMENT, "EarningsNerd")


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="left" if col == 1 else "right")


def _has_derived(dataset: dict[str, Any]) -> bool:
    return any(
        point.get("derived")
        for series in dataset.get("series", [])
        for point in series.get("points", [])
    )


def _build_overview(ws: Worksheet, dataset: dict[str, Any], exported_at: datetime) -> None:
    ws.sheet_properties.tabColor = _TAB_COLOR
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100

    if _MARK_PNG.exists():
        image = XlsxImage(str(_MARK_PNG))
        # Explicit display size (the PNG is 2× for crispness; openpyxl anchors at 96dpi pixels).
        image.width, image.height = 128, 99
        ws.add_image(image, "A1")

    mode_label = "Annual" if dataset.get("mode") == "annual" else "Quarterly"
    periods = dataset.get("periods", [])

    title = ws.cell(row=7, column=1, value=f"{dataset.get('company_name', '')} — Multi-Period Analysis")
    title.font = Font(bold=True, size=15, color=PALETTE["ink"].lstrip("#"))
    subtitle = ws.cell(row=8, column=1, value="Chart & metrics data export · EarningsNerd — AI-Powered SEC Filing Analysis")
    subtitle.font = _FOOTNOTE_FONT

    facts: list[tuple[str, str]] = [
        ("Ticker", str(dataset.get("ticker", ""))),
        ("Mode", mode_label),
        ("Periods", f"{dataset.get('period_key', '')} ({len(periods)} periods)"),
        ("Exported", exported_at.strftime("%B %d, %Y")),
        ("Source", "SEC XBRL (companyfacts) via EarningsNerd — earningsnerd.io"),
    ]
    row = 10
    for label, value in facts:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = _INK_FONT
        row += 1

    notes = [
        "Percent series (margins, growth) are true Excel percentages: a cell showing 55.8% stores 0.558.",
        "Monetary values are raw full-precision dollars — apply your own scaling as needed.",
    ]
    if _has_derived(dataset):
        notes.append(
            "Cells with a comment marker are computed Q4 estimates (full year minus reported "
            "year-to-date quarters; EPS shares-based) — shown as † in the product."
        )
    row += 1
    for note in notes:
        ws.cell(row=row, column=1, value=note).font = _FOOTNOTE_FONT
        row += 1

    row += 1
    disclaimer = ws.cell(
        row=row,
        column=1,
        value=(
            "This export is for general informational and research purposes only. It is not "
            "investment, financial, legal, accounting, or tax advice, and not a recommendation to "
            "buy, sell, or hold any security. All underlying figures come from the company's XBRL "
            "filings with the U.S. SEC (EDGAR companyfacts); growth rates, margins, ratios, and "
            "computed-Q4 values are calculated by EarningsNerd and do not appear in the filings "
            "themselves. Use is subject to the EarningsNerd Terms of Service "
            "(earningsnerd.io/terms). EarningsNerd is not affiliated with or endorsed by the SEC."
        ),
    )
    disclaimer.font = _FOOTNOTE_FONT
    disclaimer.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 78


def _build_metrics(ws: Worksheet, dataset: dict[str, Any]) -> None:
    """The full grid: one row per series, one typed column per period (+ annual window figures)."""
    ws.sheet_properties.tabColor = _TAB_COLOR
    periods = [p.get("key", "") for p in dataset.get("periods", [])]
    series_list = dataset.get("series", [])
    # Window figures exist only in annual mode (CAGR / window pp) — omit the columns otherwise.
    has_window = any(
        s.get("cagr") is not None or s.get("window_pp") is not None for s in series_list
    )

    headers = ["Metric", "Concept", "Unit", *periods]
    if has_window:
        headers += ["Window growth", "Window"]
    _write_header_row(ws, headers)

    for row_index, series in enumerate(series_list, start=2):
        unit = series.get("unit", "USD")
        percent = bool(series.get("percent"))
        label_cell = ws.cell(row=row_index, column=1, value=series.get("label", ""))
        label_cell.font = _INK_FONT
        label_cell.border = _CELL_BORDER
        ws.cell(row=row_index, column=2, value=series.get("concept", "")).border = _CELL_BORDER
        unit_cell = ws.cell(row=row_index, column=3, value="percent" if percent else str(unit))
        unit_cell.border = _CELL_BORDER

        by_period = {p.get("period"): p for p in series.get("points", [])}
        for offset, period in enumerate(periods):
            _write_point(ws, row_index, 4 + offset, by_period.get(period) or {}, unit, percent)

        if has_window:
            growth_cell = ws.cell(row=row_index, column=4 + len(periods))
            growth_cell.border = _CELL_BORDER
            window_cell = ws.cell(row=row_index, column=5 + len(periods))
            window_cell.border = _CELL_BORDER
            if percent and series.get("window_pp") is not None:
                growth_cell.value = series["window_pp"]
                growth_cell.number_format = '0.0" pp"'
                window_cell.value = series.get("window_pp_range") or ""
            elif series.get("cagr") is not None:
                growth_cell.value = series["cagr"]
                growth_cell.number_format = "0.0%"
                window_cell.value = series.get("cagr_window") or ""

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 11
    for offset in range(len(periods)):
        ws.column_dimensions[get_column_letter(4 + offset)].width = 14
    if has_window:
        ws.column_dimensions[get_column_letter(4 + len(periods))].width = 14
        ws.column_dimensions[get_column_letter(5 + len(periods))].width = 17


def _style_line_series(chart: LineChart) -> None:
    for index, chart_series in enumerate(chart.series):
        color = CHART_SERIES[index % len(CHART_SERIES)]
        chart_series.graphicalProperties.line.solidFill = color
        chart_series.graphicalProperties.line.width = 22000  # EMU ≈ 1.75pt, the on-page stroke
        chart_series.smooth = False


def _anchor_for(ws: Worksheet, header_count: int) -> str:
    return f"{get_column_letter(header_count + 2)}2"


def _add_top_line_sheet(wb: Workbook, dataset: dict[str, Any]) -> None:
    """The bar panel: top line (revenue, else net interest income) + YoY growth as a column."""
    by_concept = {s.get("concept"): s for s in dataset.get("series", [])}
    top = by_concept.get("revenue") or by_concept.get("net_interest_income")
    if top is None:
        return
    ws = wb.create_sheet(_sheet_title(f"{top.get('label', 'Top line')} & growth"))
    ws.sheet_properties.tabColor = _TAB_COLOR

    unit = top.get("unit", "USD")
    headers = ["Period", str(top.get("label", "Top line")), "YoY growth"]
    _write_header_row(ws, headers)
    by_period = {p.get("period"): p for p in top.get("points", [])}
    periods = [p.get("key", "") for p in dataset.get("periods", [])]
    for row_index, period in enumerate(periods, start=2):
        period_cell = ws.cell(row=row_index, column=1, value=period)
        period_cell.border = _CELL_BORDER
        point = by_period.get(period) or {}
        _write_point(ws, row_index, 2, point, unit, percent=False)
        growth_cell = ws.cell(row=row_index, column=3)
        growth_cell.border = _CELL_BORDER
        yoy = point.get("yoy")
        if isinstance(yoy, (int, float)):
            growth_cell.value = yoy
            growth_cell.number_format = "0.0%"
        elif yoy is not None:
            growth_cell.value = "n/m"  # sign-flip sentinel — not a meaningful percentage

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    chart = BarChart()
    chart.type = "col"
    chart.title = str(top.get("label", "Top line"))
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(periods))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + len(periods)))
    chart.series[0].graphicalProperties.solidFill = CHART_SERIES[0]
    chart.width, chart.height = 20, 10
    ws.add_chart(chart, _anchor_for(ws, len(headers)))


def _add_line_panel_sheets(wb: Workbook, dataset: dict[str, Any]) -> None:
    by_concept = {s.get("concept"): s for s in dataset.get("series", [])}
    periods = [p.get("key", "") for p in dataset.get("periods", [])]

    for title, concepts in _LINE_PANELS:
        present = [(concept, label) for concept, label in concepts if concept in by_concept]
        if not present:
            continue
        ws = wb.create_sheet(_sheet_title(title))
        ws.sheet_properties.tabColor = _TAB_COLOR

        headers = ["Period", *(label for _, label in present)]
        _write_header_row(ws, headers)
        maps = [
            (
                {p.get("period"): p for p in by_concept[concept].get("points", [])},
                by_concept[concept].get("unit", "USD"),
                bool(by_concept[concept].get("percent")),
            )
            for concept, _ in present
        ]
        for row_index, period in enumerate(periods, start=2):
            period_cell = ws.cell(row=row_index, column=1, value=period)
            period_cell.border = _CELL_BORDER
            for col_offset, (by_period, unit, percent) in enumerate(maps):
                _write_point(ws, row_index, 2 + col_offset, by_period.get(period) or {}, unit, percent)

        ws.column_dimensions["A"].width = 12
        for col_offset in range(len(present)):
            ws.column_dimensions[get_column_letter(2 + col_offset)].width = 16

        chart = LineChart()
        chart.title = title
        data = Reference(ws, min_col=2, max_col=1 + len(present), min_row=1, max_row=1 + len(periods))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + len(periods)))
        _style_line_series(chart)
        chart.width, chart.height = 20, 10
        ws.add_chart(chart, _anchor_for(ws, len(headers)))


def build_analysis_workbook(
    dataset: dict[str, Any], exported_at: Optional[datetime] = None
) -> bytes:
    """The branded analysis workbook as xlsx bytes (pure — tests read them straight back).

    ``exported_at`` is injectable for deterministic tests; the route leaves it defaulted.
    """
    exported = exported_at or utcnow()
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    _build_overview(overview, dataset, exported)
    _build_metrics(workbook.create_sheet("Metrics"), dataset)
    _add_top_line_sheet(workbook, dataset)
    _add_line_panel_sheets(workbook, dataset)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
